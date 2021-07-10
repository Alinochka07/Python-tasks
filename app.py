from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, redirect


app = Flask(__name__)


@app.route('/')
def my_project():
    wb = Workbook()
    ws = wb.active
    wb.save('mylist.xlsx')
    page = wb.active
    page['A1'] = 'Task47 My list'
    page['A2'] = 'Hello list'
    wb.save('mylist.xlsx')
    return render_template('index.html')



# @app.route('/')
# def myproject():
#     e = Workbook()
#     e.save('myexcel.xlsx')
#     return render_template('index.html')
